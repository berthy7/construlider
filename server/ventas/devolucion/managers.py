from ...operaciones.bitacora.managers import *

from server.common.managers import SuperManager
from server.ventas.reserva.models import *
from sqlalchemy.sql import func,or_,and_
from server.terrenos.terreno.managers import *
from datetime import datetime
from openpyxl import load_workbook


class DevolucionManager(SuperManager):
    def __init__(self, db):
        super().__init__(Reserva, db)

    def list_all(self):
        return dict(objects=self.db.query(self.entity).join(Estadoreserva).filter(Estadoreserva.nombre == "Cancelado"))


    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.enabled == True).order_by(self.entity.nombre.asc()).all()

    def insert(self, dictionary):
        TerrenoManager(self.db).cambiar_estado(dictionary['fkterreno'], False)
        dictionary['fechaDevolucion'] = datetime.strptime(dictionary['fechaDevolucion'], '%d/%m/%Y')
        dictionary['tipo'] = "Devolucion"
        objeto = DevolucionManager(self.db).entity(**dictionary)
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registro Devolucion.", fecha=fecha, tabla="reserva",
                     identificador=a.id)
        super().insert(b)


        return a

    def update(self, dictionary):
        objeto = DevolucionManager(self.db).entity(**dictionary)
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modifico Devolucion.", fecha=fecha,tabla="reserva", identificador=a.id)
        super().insert(b)

        return a

    def cancel(self, id, Usuario, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).first()
        TerrenoManager(self.db).cambiar_estado(x.fkterreno, True)
        estado = EstadoreservaManager(self.db).obtener_x_nombre("Cancelado")

        fecha = BitacoraManager(self.db).fecha_actual()

        x.fkestadoreserva = estado.id
        x.fechaDevolucion = fecha
        x.tipo = ""

        b = Bitacora(fkusuario=Usuario, ip=ip, accion="Cancelo Devolucion", fecha=fecha)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()

        return x

    def delete(self, id, Usuario, ip,enable):
        x = self.db.query(self.entity).filter(self.entity.id == id).first()

        x.enabled = enable

        if enable:
            mensaje = "Se habilitó reserva."
        else:
            mensaje = "Se deshabilitó reserva."

        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=Usuario, ip=ip, accion=mensaje, fecha=fecha,tabla="reserva", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()

        return mensaje

    def reporte_excel(self):
        fecha = datetime.now()
        cname = "Devolucion_" + fecha.strftime('%Y-%m-%d') + ".xlsx"

        wb = Workbook()

        ws = wb.active
        ws.title = 'Reporte Devolucion'

        indice = 0
        # --------------------------------------------------------------------
        indice = indice + 1
        ws['A' + str(indice)] = 'PERSONAL QUE NO CUENTA CON LA TOTALIDAD DE DIAS DE VACACION REQUERIDO.'
        indice = indice + 2

        ws['A' + str(indice)] = 'Nº'
        ws['B' + str(indice)] = 'CODIGO'
        ws['C' + str(indice)] = 'NOMBRE COMPLETO'
        ws['D' + str(indice)] = 'DIAS DISPONIBLES'


        wb.save("server/common/resources/downloads/" + cname)

        return cname

class EstadoreservaManager(SuperManager):
    def __init__(self, db):
        super().__init__(Estadoreserva, db)

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.enabled == True).all()

    def obtener_x_nombre(self,nombre):
        return self.db.query(self.entity).filter(self.entity.nombre == nombre).filter(self.entity.enabled == True).first()



